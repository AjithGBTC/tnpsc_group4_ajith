import io
import uuid
from pathlib import Path
from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, Query, UploadFile, status
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from app.database.session import get_db
from app.dependencies.auth import current_user, require_permissions
from app.models.entities import Question, Taxonomy
from app.schemas.content import QuestionCreate, TaxonomyCreate
from app.services.notifications import send_topic_notification

router = APIRouter(tags=["Content"])

@router.get("/taxonomies/{kind}")
async def list_taxonomy(kind: str, page: int = Query(1, ge=1), size: int = Query(50, ge=1, le=100), db: AsyncSession = Depends(get_db)):
    query = select(Taxonomy).where(Taxonomy.kind == kind, Taxonomy.deleted_at.is_(None)).order_by(Taxonomy.sequence, Taxonomy.name).offset((page - 1) * size).limit(size)
    items = (await db.scalars(query)).all()
    return {"data": [{"id": str(x.id), "name": x.name, "parent_id": str(x.parent_id) if x.parent_id else None, "status": x.status} for x in items], "meta": {"page": page, "size": size}}

@router.post("/taxonomies/{kind}", status_code=status.HTTP_201_CREATED, dependencies=[Depends(require_permissions("content:write"))])
async def create_taxonomy(kind: str, payload: TaxonomyCreate, user=Depends(current_user), db: AsyncSession = Depends(get_db)):
    entity = Taxonomy(kind=kind, **payload.model_dump(), created_by=user.id, updated_by=user.id)
    db.add(entity); await db.commit(); await db.refresh(entity)
    return {"data": {"id": str(entity.id), "name": entity.name}}

@router.get("/questions")
async def list_questions(exam_id: str | None = None, approval_status: str | None = None, page: int = Query(1, ge=1), size: int = Query(50, ge=1, le=100), db: AsyncSession = Depends(get_db)):
    query = select(Question).where(Question.deleted_at.is_(None))
    if exam_id: query = query.where(Question.exam_id == exam_id)
    if approval_status: query = query.where(Question.approval_status == approval_status)
    items = (await db.scalars(query.order_by(Question.created_at.desc()).offset((page - 1) * size).limit(size))).all()
    return {"data": [{"id": str(q.id), "body": q.body, "status": q.approval_status, "difficulty": q.difficulty} for q in items], "meta": {"page": page, "size": size}}

@router.post("/questions", status_code=status.HTTP_201_CREATED, dependencies=[Depends(require_permissions("questions:write"))])
async def create_question(payload: QuestionCreate, background_tasks: BackgroundTasks, user=Depends(current_user), db: AsyncSession = Depends(get_db)):
    question = Question(**payload.model_dump(), created_by=user.id, updated_by=user.id)
    db.add(question); await db.commit(); await db.refresh(question)
    if question.approval_status == "approved":
        background_tasks.add_task(send_topic_notification, "New question added", "A new practice question is available", "question", str(question.id))
    return {"data": {"id": str(question.id), "approval_status": question.approval_status}}

@router.put("/questions/{question_id:uuid}", dependencies=[Depends(require_permissions("questions:write"))])
async def update_question(question_id: uuid.UUID, payload: QuestionCreate, user=Depends(current_user), db: AsyncSession = Depends(get_db)):
    question = await db.get(Question, question_id)
    if not question or question.deleted_at: raise HTTPException(404, "Question not found")
    for key, value in payload.model_dump().items(): setattr(question, key, value)
    question.updated_by = user.id; await db.commit(); return {"message": "Question updated"}

@router.delete("/questions/{question_id:uuid}", status_code=204, dependencies=[Depends(require_permissions("questions:write"))])
async def delete_question(question_id: uuid.UUID, user=Depends(current_user), db: AsyncSession = Depends(get_db)) -> None:
    question = await db.get(Question, question_id)
    if not question or question.deleted_at: raise HTTPException(404, "Question not found")
    from datetime import UTC, datetime
    question.deleted_at, question.updated_by = datetime.now(UTC), user.id; await db.commit()

@router.get("/questions/export")
async def export_questions(db: AsyncSession = Depends(get_db), _=Depends(require_permissions("questions:write"))):
    import json
    wb = Workbook(); ws = wb.active; ws.title = "Questions"
    ws.append(["exam_id", "subject_id", "topic_id", "body", "options_json", "answer_json", "explanation", "difficulty", "question_type", "marks", "negative_marks", "approval_status"])
    rows = (await db.scalars(select(Question).where(Question.deleted_at.is_(None)))).all()
    for q in rows: ws.append([str(q.exam_id), str(q.subject_id or ""), str(q.topic_id or ""), q.body, json.dumps(q.options), json.dumps(q.answer), q.explanation, q.difficulty, q.question_type, q.marks, q.negative_marks, q.approval_status])
    export_dir = Path("uploads/exports")
    export_dir.mkdir(parents=True, exist_ok=True)
    filename = f"questions-{uuid.uuid4().hex}.xlsx"
    wb.save(export_dir / filename)
    return {"download_url": f"/uploads/exports/{filename}"}

@router.post("/questions/import", dependencies=[Depends(require_permissions("questions:write"))])
async def import_questions(file: UploadFile = File(...), user=Depends(current_user), db: AsyncSession = Depends(get_db)):
    import json
    if not file.filename or not file.filename.lower().endswith(".xlsx"): raise HTTPException(400, "Upload an .xlsx file")
    try: rows = list(load_workbook(io.BytesIO(await file.read()), read_only=True, data_only=True).active.iter_rows(min_row=2, values_only=True))
    except Exception as exc: raise HTTPException(400, "Invalid Excel workbook") from exc
    created = 0
    for index, row in enumerate(rows, 2):
        if not row or not row[0] or not row[3]: continue
        try:
            db.add(Question(exam_id=uuid.UUID(str(row[0])), subject_id=uuid.UUID(str(row[1])) if row[1] else None, topic_id=uuid.UUID(str(row[2])) if row[2] else None, body=str(row[3]), options=json.loads(row[4]), answer=json.loads(row[5]), explanation=row[6], difficulty=row[7] or "medium", question_type=row[8] or "single_choice", marks=int(row[9] or 1), negative_marks=int(row[10] or 0), approval_status=row[11] or "draft", created_by=user.id, updated_by=user.id)); created += 1
        except Exception as exc: raise HTTPException(422, f"Invalid row {index}: {exc}") from exc
    await db.commit(); return {"data": {"created": created}}
